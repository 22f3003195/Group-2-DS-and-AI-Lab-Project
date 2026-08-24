"""Build a normalised reference-range database from lab_test_reference_ranges_1.xlsx.

Run:
    python -m code.reference_ranges.build_reference_db            # writes reference_db.json
    python -m code.reference_ranges.build_reference_db --report   # + coverage report

The output JSON is the single source of truth for the pipeline. Every numeric
bound is parsed here, once, at build time. Nothing downstream re-parses a range
string and nothing downstream asks a language model what "normal" means.

Two rules govern the parse:

1. The Notes column wins. The workbook's own Read Me states that qualitative
   analytes, body-fluid variants and clinically-interpreted tests have no fixed
   normal range. Where a note says so, the row is marked non-evaluable even if
   the Reference Range cell contains digits. This matters: `Anti-Nuclear
   Antibody` reads "Negative (<1:40)" and `Length of Urine Collection` reads
   "24" - both would otherwise be parsed into comparable numbers.

2. Anything the grammar cannot parse exactly becomes non-evaluable with a
   recorded reason. There is no partial credit and no fallback guess.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:  # allow both `python -m code.reference_ranges.build_reference_db` and direct run
    from .units import normalize_unit, unit_group
except ImportError:  # pragma: no cover
    from units import normalize_unit, unit_group  # type: ignore

_HERE = Path(__file__).resolve().parent
# The workbook ships next to this module so the Space never needs the repo root.
DEFAULT_XLSX = _HERE / "lab_test_reference_ranges_1.xlsx"
if not DEFAULT_XLSX.exists():  # fall back to a repo-root copy for local rebuilds
    DEFAULT_XLSX = _HERE.parents[2] / "lab_test_reference_ranges_1.xlsx"
DEFAULT_OUT = _HERE / "reference_db.json"
# Rows the vendor workbook lacks, kept in a reviewable JSON file rather than
# edited into the binary .xlsx so additions show up in a git diff with their
# sources attached. See its _readme for the sourcing policy.
DEFAULT_SUPPLEMENT = _HERE / "additional_ranges.json"

SHEET = "Reference Ranges"
SCHEMA_VERSION = 1

# --------------------------------------------------------------------------
# Note policy - exhaustive over the 12 distinct Notes values in the workbook
# --------------------------------------------------------------------------
# BLOCK      -> row can never be numerically classified
# CAVEAT     -> row is usable, but the note is a precondition the patient must see
# PROVENANCE -> bookkeeping only, no effect on classification

NOTE_POLICY: list[tuple[str, str, str]] = [
    # (substring to match, policy, reason code)
    ("Qualitative / not a numeric analyte", "BLOCK", "QUALITATIVE_ANALYTE"),
    ("Body-fluid values interpreted relative to serum", "BLOCK", "BODY_FLUID_NO_FIXED_RANGE"),
    ("No standardized reference range", "BLOCK", "NO_STANDARDIZED_RANGE"),
    ("age-specific nomogram", "BLOCK", "NOMOGRAM_REQUIRED"),
    ("time-dependent", "BLOCK", "TIME_DEPENDENT_THRESHOLD"),
    ("Pattern-based", "BLOCK", "PATTERN_BASED"),
    ("Fasting", "CAVEAT", "REQUIRES_FASTING"),
    ("Ionized", "CAVEAT", "IONIZED_FRACTION"),
    ("Non-pregnant reference only", "CAVEAT", "NON_PREGNANT_ONLY"),
    ("Source:", "PROVENANCE", ""),
    ("Brand name alias", "PROVENANCE", ""),
]


def classify_note(note: str | None) -> tuple[str, str, str]:
    """Return (policy, reason_code, note_text) for a Notes cell."""
    if note is None or not str(note).strip():
        return "NONE", "", ""
    text = str(note).strip()
    for needle, policy, reason in NOTE_POLICY:
        if needle.lower() in text.lower():
            return policy, reason, text
    # Unrecognised note: be conservative but do not block - record it verbatim
    # so a human can extend NOTE_POLICY.
    return "UNKNOWN_NOTE", "", text


# --------------------------------------------------------------------------
# Range grammar
# --------------------------------------------------------------------------

DASH = "–—−-"  # en dash, em dash, minus sign, hyphen
NUM = r"\d+(?:\.\d+)?"
_SEP = rf"[{DASH}]"

# The workbook never uses a hyphen as a range separator (verified: hyphens only
# occur inside words like "non-diabetic"), but reports might, so both are read.
RE_PLAIN = re.compile(rf"^({NUM})\s*{_SEP}\s*({NUM})$")
RE_BOUND = re.compile(rf"^([<>]=?|≤|≥)\s*({NUM})$")
RE_SINGLE = re.compile(rf"^({NUM})$")
# "0-3/HPF (urine)" - unit glued into the range cell
RE_RANGE_HPF = re.compile(rf"^({NUM})\s*{_SEP}\s*({NUM})\s*/\s*(HPF|LPF)\b", re.I)
# "13.2-16.6 (M), 11.6-15.0 (F)"  /  ">40 (M), >50 (F)"
RE_SEX_RANGE = re.compile(rf"({NUM})\s*{_SEP}\s*({NUM})\s*\(\s*(M|F)\s*\)", re.I)
RE_SEX_BOUND = re.compile(rf"([<>]=?|≤|≥)\s*({NUM})\s*\(\s*(M|F)\s*\)", re.I)
# "30-200 (M 39-308, F 26-192)" - sex sub-ranges nested in the qualifier
RE_SEX_INNER = re.compile(rf"\b(M|F)\s*({NUM})\s*{_SEP}\s*({NUM})", re.I)
# "4.0-5.6 (non-diabetic)" - a range with a trailing context qualifier
RE_RANGE_QUAL = re.compile(rf"^({NUM})\s*{_SEP}\s*({NUM})\s*\((.+)\)$")
RE_BOUND_QUAL = re.compile(rf"^([<>]=?|≤|≥)\s*({NUM})\s*[\(\s](.+?)\)?$")
RE_SINGLE_QUAL = re.compile(rf"^({NUM})\s*\((.+)\)$")
# "20-30 peak / <8 trough" - therapeutic drug monitoring, needs dose timing
RE_PEAK_TROUGH = re.compile(r"\b(peak|trough)\b", re.I)
# "Rare (0-2/HPF)", "Few (0-5/HPF)" - a semi-quantitative grade whose numeric
# density is still exact. Deliberately restricted to /HPF: it must not swallow
# result-valued prefixes like "Negative (<6.5 IU/mL)", where the word IS the
# result and the number is only the assay cutoff.
RE_GRADED_HPF = re.compile(
    rf"^([A-Za-z][A-Za-z/ ]*?)\s*\(\s*({NUM}\s*{_SEP}\s*{NUM}\s*/\s*HPF)\s*\)$", re.I
)

EMPTY_RANGE = {"", "-", "–", "—", "none", "nan"}


def _mk(low, high, incl_low=True, incl_high=True) -> dict[str, Any]:
    return {
        "low": low,
        "high": high,
        "inclusive_low": incl_low,
        "inclusive_high": incl_high,
    }


def _bound_from_op(op: str, val: float) -> dict[str, Any]:
    """Turn '<150' / '>=35' into a one-sided bound."""
    op = op.replace("≤", "<=").replace("≥", ">=")
    if op in ("<", "<="):
        return _mk(None, val, True, op == "<=")
    return _mk(val, None, op == ">=", True)


def parse_range(raw: str | None) -> dict[str, Any]:
    """Parse a Reference Range cell into structured bounds.

    Returns a dict with:
        kind    - range | bound | expected_zero | sex_specific | non_numeric
        bounds  - {context: bound}, context is "any", "M" or "F"
        qualifier, reason
    """
    fail = lambda reason: {  # noqa: E731
        "kind": "non_numeric",
        "bounds": {},
        "qualifier": None,
        "reason": reason,
    }

    if raw is None:
        return fail("NO_RANGE_GIVEN")
    text = unicodedata.normalize("NFKC", str(raw)).strip()
    if text.lower() in EMPTY_RANGE:
        return fail("NO_RANGE_GIVEN")

    # Therapeutic drug monitoring: a peak and a trough target cannot be applied
    # without knowing when the sample was drawn relative to the dose.
    if RE_PEAK_TROUGH.search(text):
        return fail("AMBIGUOUS_PEAK_TROUGH")

    # --- sex-specific forms (checked first: they contain plain ranges) -------
    bounds: dict[str, Any] = {}
    for lo, hi, sx in RE_SEX_RANGE.findall(text):
        bounds[sx.upper()] = _mk(float(lo), float(hi))
    for op, val, sx in RE_SEX_BOUND.findall(text):
        bounds[sx.upper()] = _bound_from_op(op, float(val))

    if not bounds:
        # nested form: "30-200 (M 39-308, F 26-192)"
        m_outer = RE_RANGE_QUAL.match(text)
        if m_outer:
            inner = RE_SEX_INNER.findall(m_outer.group(3))
            if inner:
                for sx, lo, hi in inner:
                    bounds[sx.upper()] = _mk(float(lo), float(hi))
                if bounds:
                    bounds["any"] = _mk(float(m_outer.group(1)), float(m_outer.group(2)))

    if bounds:
        if "any" not in bounds:
            lows = [b["low"] for b in bounds.values() if b["low"] is not None]
            highs = [b["high"] for b in bounds.values() if b["high"] is not None]
            complete = len(bounds) >= 2 and {"M", "F"} <= set(bounds)
            if complete:
                bounds["any"] = _mk(
                    min(lows) if lows else None,
                    max(highs) if highs else None,
                )
        return {
            "kind": "sex_specific",
            "bounds": bounds,
            "qualifier": None,
            # A row giving only one sex (e.g. "72-235 (M)") cannot be applied to
            # another patient, so no union is synthesised for it.
            "reason": "" if "any" in bounds else "SEX_SPECIFIC_INCOMPLETE",
        }

    # --- graded microscopy "Rare (0-2/HPF)" ---------------------------------
    m = RE_GRADED_HPF.match(text)
    if m:
        inner = parse_range(m.group(2))
        if inner["kind"] == "range":
            inner["qualifier"] = m.group(1).strip()
            return inner
        return fail("UNPARSEABLE_RANGE")

    # --- microscopy "0-3/HPF (urine)" ---------------------------------------
    m = RE_RANGE_HPF.match(text)
    if m:
        if m.group(3).upper() == "LPF":
            return fail("UNPARSEABLE_RANGE")
        qual = text[m.end():].strip(" ()") or None
        return {
            "kind": "range",
            "bounds": {"any": _mk(float(m.group(1)), float(m.group(2)))},
            "qualifier": qual,
            "reason": "",
        }

    # --- plain two-sided range ----------------------------------------------
    m = RE_PLAIN.match(text)
    if m:
        return {
            "kind": "range",
            "bounds": {"any": _mk(float(m.group(1)), float(m.group(2)))},
            "qualifier": None,
            "reason": "",
        }

    # --- range with a trailing qualifier ------------------------------------
    m = RE_RANGE_QUAL.match(text)
    if m:
        return {
            "kind": "range",
            "bounds": {"any": _mk(float(m.group(1)), float(m.group(2)))},
            "qualifier": m.group(3).strip(),
            "reason": "",
        }

    # --- one-sided bound ----------------------------------------------------
    m = RE_BOUND.match(text)
    if m:
        return {
            "kind": "bound",
            "bounds": {"any": _bound_from_op(m.group(1), float(m.group(2)))},
            "qualifier": None,
            "reason": "",
        }

    m = RE_BOUND_QUAL.match(text)
    if m and not re.search(r"\d\s*:\s*\d", text):  # exclude titres like "<1:40"
        return {
            "kind": "bound",
            "bounds": {"any": _bound_from_op(m.group(1), float(m.group(2)))},
            "qualifier": m.group(3).strip(" ()"),
            "reason": "",
        }

    # --- bare single number: "should be zero / absent" ----------------------
    m = RE_SINGLE.match(text) or RE_SINGLE_QUAL.match(text)
    if m:
        val = float(m.group(1))
        qual = m.group(2).strip() if m.lastindex and m.lastindex > 1 else None
        if val == 0:
            return {
                "kind": "expected_zero",
                "bounds": {"any": _mk(0.0, 0.0)},
                "qualifier": qual,
                "reason": "",
            }
        # A lone non-zero number is not a range; refuse rather than invent one.
        return fail("SINGLE_VALUE_NOT_A_RANGE")

    if not re.search(r"\d", text):
        return fail("QUALITATIVE_TEXT")
    return fail("UNPARSEABLE_RANGE")


# --------------------------------------------------------------------------
# Test-name normalisation
# --------------------------------------------------------------------------

def normalize_name(raw: str | None) -> str:
    """Fold a test name for exact/alias matching.

    Handles the workbook's own noise: wrapper punctuation ("(Albumin)",
    "<Albumin>"), doubled spaces, and case.

    The trailing '#' is NOT discarded. In this workbook it is the absolute-count
    marker, and it is the only thing separating `Lymphs` (20-40 %) from
    `Lymphs#` (1.0-4.8 x10^3/uL). Folding them together would recreate exactly
    the percentage-vs-absolute-count confusion this module exists to remove, so
    '#' is preserved as its own token.
    """
    if raw is None:
        return ""
    s = unicodedata.normalize("NFKC", str(raw)).strip()
    # Strip wrapper punctuation only when it WRAPS the whole name, as in the
    # workbook's "(Albumin)" and "<Albumin>" junk rows. A blind strip also ate
    # the closing bracket of "Thyroxine (T4)" and "SGPT (ALT)", leaving an
    # unbalanced "sgpt (alt" that no synonym logic could split.
    while len(s) > 1 and ((s[0], s[-1]) in (("(", ")"), ("<", ">"),
                                            ("[", "]"), ("{", "}"))):
        s = s[1:-1].strip()
    s = s.replace("–", "-").replace("—", "-")
    s = s.lower()
    s = re.sub(r"\*+", " ", s)
    s = re.sub(r"\s*#\s*", " #", s)
    s = re.sub(r"[^a-z0-9%/+.,()#-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\s*,\s*", ", ", s)
    return s


def _singularize(token: str) -> str:
    """Crude English singulariser, sufficient for lab test nomenclature.

    Reports say "ABSOLUTE LYMPHOCYTES"; the table says "Absolute Lymphocyte
    Count". Without this, the shared token count drops and plain string
    similarity ranks the bare substring `Lymphocytes` (20-40 %) above the
    correct `Absolute Lymphocyte Count` (1.0-4.8 x10^3/uL).
    """
    if len(token) > 4 and token.endswith("s") and not token.endswith(("ss", "us", "is")):
        return token[:-1]
    return token


def name_tokens(normalized: str) -> str:
    """Punctuation-free, singularised token string for fuzzy matching."""
    raw = re.sub(r"[^a-z0-9]+", " ", normalized).split()
    return " ".join(_singularize(t) for t in raw)


# --------------------------------------------------------------------------
# Build
# --------------------------------------------------------------------------

def load_supplement(path: Path) -> list[tuple]:
    """Read additional_ranges.json as (name, range, unit, note) tuples."""
    if not path.exists():
        return []
    doc = json.loads(path.read_text(encoding="utf-8"))
    return [
        (r["Test Name"], r["Reference Range"], r["Units"], r["Notes"])
        for r in doc.get("rows", [])
    ]


def build(xlsx_path: Path, supplement_path: Path | None = DEFAULT_SUPPLEMENT) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET]

    disclaimer = [
        " ".join(str(c) for c in row if c is not None).strip()
        for row in wb["Read Me"].iter_rows(values_only=True)
        if any(c is not None for c in row)
    ]

    source_rows = list(ws.iter_rows(min_row=2, values_only=True))
    supplement = load_supplement(supplement_path) if supplement_path else []
    # A supplementary row never silently overwrites a workbook row.
    existing = {normalize_name(r[0]) for r in source_rows if r[0]}
    added, skipped = [], []
    for row in supplement:
        (added if normalize_name(row[0]) not in existing else skipped).append(row)
    if skipped:
        print(f"note: {len(skipped)} supplementary row(s) already in the workbook, "
              f"kept the workbook's: {[r[0] for r in skipped]}")

    entries: list[dict[str, Any]] = []
    for idx, (name, rng, unit, note) in enumerate(source_rows + added):
        if name is None or not str(name).strip():
            continue

        policy, note_reason, note_text = classify_note(note)
        parsed = parse_range(rng)

        evaluable = parsed["kind"] != "non_numeric"
        reason = parsed["reason"]

        # Rule 1: the Notes column overrides a successful parse.
        if policy == "BLOCK":
            evaluable = False
            reason = note_reason
        elif not evaluable and not reason:
            reason = "UNPARSEABLE_RANGE"

        # A sex-specific row with only one sex given stays evaluable, but only
        # for that sex; the resolver enforces this via the missing "any" key.
        if parsed["kind"] == "sex_specific" and reason == "SEX_SPECIFIC_INCOMPLETE":
            evaluable = True

        norm = normalize_name(name)
        entries.append(
            {
                "id": idx,
                "test_name": str(name).strip(),
                "normalized": norm,
                "tokens": name_tokens(norm),
                "raw_range": None if rng is None else str(rng).strip(),
                "unit_raw": None if unit is None else str(unit).strip(),
                "unit_normalized": normalize_unit(unit),
                "unit_group": unit_group(unit),
                "note": note_text or None,
                "note_policy": policy,
                "kind": parsed["kind"],
                "bounds": parsed["bounds"],
                "qualifier": parsed["qualifier"],
                "evaluable": evaluable,
                "not_evaluable_reason": None if evaluable else (reason or "UNPARSEABLE_RANGE"),
                "caveat": note_text if policy == "CAVEAT" else None,
            }
        )

    digest = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()
    return {
        "schema_version": SCHEMA_VERSION,
        "source_file": xlsx_path.name,
        "source_sha256": digest,
        "built_at": _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds"),
        "row_count": len(entries),
        "supplement_rows": len(added),
        "disclaimer": disclaimer,
        "entries": entries,
    }


def report(db: dict[str, Any]) -> str:
    entries = db["entries"]
    out: list[str] = []
    out.append(f"rows: {len(entries)}")
    ev = [e for e in entries if e["evaluable"]]
    out.append(f"evaluable: {len(ev)}  ({len(ev)/len(entries):.1%})")
    out.append("")
    out.append("by kind:")
    for k, v in Counter(e["kind"] for e in entries).most_common():
        out.append(f"  {k:16s} {v}")
    out.append("")
    out.append("non-evaluable reasons:")
    for k, v in Counter(
        e["not_evaluable_reason"] for e in entries if not e["evaluable"]
    ).most_common():
        out.append(f"  {str(k):28s} {v}")
    out.append("")
    out.append("note policy:")
    for k, v in Counter(e["note_policy"] for e in entries).most_common():
        out.append(f"  {k:16s} {v}")
    unknown_units = Counter(
        e["unit_raw"] for e in entries if e["unit_raw"] and e["unit_group"] is None
    )
    out.append("")
    out.append(f"unrecognised units: {len(unknown_units)}")
    for k, v in unknown_units.most_common():
        out.append(f"  {k:28s} {v}")
    return "\n".join(out)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    ap.add_argument("--report", action="store_true", help="print a coverage report")
    args = ap.parse_args()

    db = build(args.xlsx)
    args.out.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {args.out}  ({db['row_count']} rows)")
    if args.report:
        print()
        print(report(db))


if __name__ == "__main__":
    main()
